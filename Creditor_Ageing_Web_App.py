# ============================================================
# TUAREG MARKETING PVT LTD - CREDITOR AGEING WEB APPLICATION
# ============================================================
# Developed for SAP Outstanding Creditors Ageing Report
# Upload SAP OS Excel + Supplier Master Excel, then generate:
# - Clean SAP OS data
# - Supplier ageing report
# - Group-wise report
# - Operative / Non-Operative report
# - Missing supplier list
# - Downloadable Excel report
# ============================================================

import io
import re
from datetime import datetime, date
from decimal import Decimal, getcontext
from pathlib import Path

import pandas as pd
import streamlit as st
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

getcontext().prec = 60

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="Creditor Ageing | Tuareg Marketing Pvt Ltd",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# PREMIUM CSS
# ============================================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .stApp {
        background:
            radial-gradient(circle at top left, rgba(255,153,51,0.18), transparent 32%),
            radial-gradient(circle at top right, rgba(19,136,8,0.13), transparent 30%),
            linear-gradient(135deg, #f8fafc 0%, #eef2ff 48%, #fff7ed 100%);
    }

    .main-header {
        background: linear-gradient(135deg, #111827 0%, #1e3a8a 42%, #065f46 100%);
        border-radius: 26px;
        padding: 26px 30px;
        margin: 10px 0 22px 0;
        color: white;
        box-shadow: 0 18px 45px rgba(15, 23, 42, 0.26);
        border: 1px solid rgba(255,255,255,0.18);
        position: relative;
        overflow: hidden;
    }

    .main-header::after {
        content: "";
        position: absolute;
        top: -70px;
        right: -70px;
        width: 220px;
        height: 220px;
        border-radius: 50%;
        background: rgba(255,255,255,0.10);
    }

    .company-title {
        font-size: 34px;
        font-weight: 800;
        letter-spacing: -0.8px;
        margin-bottom: 6px;
    }

    .sub-title {
        font-size: 16px;
        opacity: 0.92;
        font-weight: 500;
    }

    .live-clock {
        display: inline-block;
        margin-top: 14px;
        padding: 9px 14px;
        border-radius: 999px;
        background: rgba(255,255,255,0.14);
        border: 1px solid rgba(255,255,255,0.22);
        font-size: 14px;
        font-weight: 600;
    }

    .metric-card {
        background: rgba(255,255,255,0.88);
        border: 1px solid rgba(148,163,184,0.34);
        box-shadow: 0 10px 26px rgba(15, 23, 42, 0.08);
        border-radius: 22px;
        padding: 18px 20px;
        min-height: 118px;
    }

    .metric-label {
        color: #64748b;
        font-size: 13px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.7px;
    }

    .metric-value {
        color: #0f172a;
        font-size: 27px;
        font-weight: 800;
        margin-top: 8px;
    }

    .metric-note {
        color: #475569;
        font-size: 12px;
        margin-top: 5px;
    }

    .section-card {
        background: rgba(255,255,255,0.90);
        border: 1px solid rgba(148,163,184,0.35);
        border-radius: 22px;
        padding: 20px;
        box-shadow: 0 10px 24px rgba(15, 23, 42, 0.07);
        margin-bottom: 18px;
    }

    .upload-box {
        background: rgba(255,255,255,0.92);
        border: 1px dashed #2563eb;
        border-radius: 22px;
        padding: 18px;
        margin-bottom: 12px;
    }

    div[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #111827 50%, #064e3b 100%);
    }

    div[data-testid="stSidebar"] * {
        color: white;
    }

    .small-success {
        background: #ecfdf5;
        border: 1px solid #86efac;
        color: #14532d;
        padding: 11px 14px;
        border-radius: 14px;
        font-weight: 600;
    }

    .small-warning {
        background: #fffbeb;
        border: 1px solid #fcd34d;
        color: #78350f;
        padding: 11px 14px;
        border-radius: 14px;
        font-weight: 600;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }

    .stTabs [data-baseweb="tab"] {
        border-radius: 999px;
        padding: 12px 18px;
        background: rgba(255,255,255,0.70);
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# HEADER WITH LIVE DATE/TIME
# ============================================================
st.markdown(
    f"""
    <div class="main-header">
        <div class="company-title">Tuareg Marketing Pvt Ltd</div>
        <div class="sub-title">SAP Creditor Ageing Report • Supplier Outstanding Dashboard • Master Based Classification</div>
        <div class="live-clock">🕒 Live Date & Time: <span id="clock"></span></div>
    </div>
    <script>
    function updateClock() {{
        const now = new Date();
        const options = {{ weekday:'short', year:'numeric', month:'short', day:'2-digit', hour:'2-digit', minute:'2-digit', second:'2-digit' }};
        document.getElementById('clock').innerHTML = now.toLocaleString('en-IN', options);
    }}
    updateClock();
    setInterval(updateClock, 1000);
    </script>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# HELPERS
# ============================================================

def normalize_text(x):
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in ("", "nan", "none", "nat"):
        return ""
    return s


def normalize_supplier_code(x):
    s = normalize_text(x)
    if not s:
        return ""
    try:
        if re.fullmatch(r"\d+(\.0+)?", s):
            return str(int(float(s)))
    except Exception:
        pass
    return s


def is_valid_supplier_code(x):
    s = normalize_supplier_code(x)
    return bool(s) and s.isdigit() and len(s) >= 4


SCI_RE = re.compile(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$")


def normalize_reference(x):
    s = normalize_text(x)
    if not s:
        return ""
    if SCI_RE.match(s):
        try:
            d = Decimal(s)
            return format(d.to_integral_value(), "f")
        except Exception:
            return s
    try:
        if re.fullmatch(r"\d+\.0+", s):
            return str(int(float(s)))
    except Exception:
        pass
    return s


def clean_amount(x):
    s = normalize_text(x)
    if not s:
        return 0.0
    trailing_minus = s.endswith("-")
    if trailing_minus:
        s = s[:-1]
    bracket_negative = s.startswith("(") and s.endswith(")")
    if bracket_negative:
        s = s[1:-1]
    s = s.replace(",", "").replace(" ", "")
    num = pd.to_numeric(s, errors="coerce")
    if pd.isna(num):
        return 0.0
    num = float(num)
    if trailing_minus or bracket_negative:
        num = -abs(num)
    return num


def parse_date_safe(x):
    if x is None or str(x).strip().lower() in ("", "nan", "none", "nat"):
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    dt = pd.to_datetime(x, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


DD_TOKEN_RE = re.compile(r"\(.*?DD\s*:\s*([^)]+)\)", re.IGNORECASE)


def due_date_from_dd_value(dd_value):
    token = normalize_text(dd_value)
    if not token:
        return None
    token_no_space = re.sub(r"\s+", "", token).upper()
    token_letters_only = re.sub(r"[^A-Z]", "", token_no_space)
    if token_letters_only == "STD":
        return None
    digits = re.sub(r"\D", "", token_no_space)
    if len(digits) != 6:
        return None
    try:
        dd = int(digits[0:2])
        mm = int(digits[2:4])
        yy = int(digits[4:6])
        return date(2000 + yy, mm, dd)
    except Exception:
        return None


def parse_real_due_dt(text_value, pmnt_date_only, dd_column_value=None):
    due_from_dd_col = due_date_from_dd_value(dd_column_value)
    if due_from_dd_col:
        return due_from_dd_col
    txt = normalize_text(text_value)
    if txt:
        m = DD_TOKEN_RE.search(txt)
        if m:
            due_from_text = due_date_from_dd_value(m.group(1))
            if due_from_text:
                return due_from_text
    return pmnt_date_only


def ageing_bucket(due_dt):
    if due_dt is None:
        return "NO DUE DATE"
    today = date.today()
    delta = (due_dt - today).days
    if delta < 0:
        return "OVERDUE"
    if delta <= 7:
        return "DUE 0-7 DAYS"
    if delta <= 15:
        return "DUE 8-15 DAYS"
    if delta <= 30:
        return "DUE 16-30 DAYS"
    if delta <= 60:
        return "DUE 31-60 DAYS"
    return "DUE 61+ DAYS"


BUCKET_ORDER = [
    "OVERDUE",
    "DUE 0-7 DAYS",
    "DUE 8-15 DAYS",
    "DUE 16-30 DAYS",
    "DUE 31-60 DAYS",
    "DUE 61+ DAYS",
    "NO DUE DATE",
]


def format_inr(x):
    try:
        x = float(x)
    except Exception:
        return "₹0"
    sign = "-" if x < 0 else ""
    x = abs(x)
    if x >= 10_000_000:
        return f"{sign}₹{x/10_000_000:,.2f} Cr"
    if x >= 100_000:
        return f"{sign}₹{x/100_000:,.2f} L"
    return f"{sign}₹{x:,.0f}"


def find_header_row_and_columns(ws):
    header_row = None
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [normalize_text(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)]
        if "supplier" in values:
            header_row = r
            break
    if not header_row:
        raise Exception("Supplier column not found in master file. Please check master header row.")
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = normalize_text(ws.cell(header_row, c).value)
        if v:
            col_map[v.lower()] = c
    return header_row, col_map


def read_master_from_upload(master_file):
    master_bytes = master_file.getvalue()
    wb = load_workbook(io.BytesIO(master_bytes))
    ws = wb.active
    header_row, col_map = find_header_row_and_columns(ws)

    sup_c = col_map.get("supplier")
    name_c = col_map.get("name") or col_map.get("vendor name") or col_map.get("supplier name")
    group_c = col_map.get("group")
    view_c = col_map.get("view - y/n") or col_map.get("view-y/n") or col_map.get("view") or col_map.get("operative")
    type_c = col_map.get("supplier type") or col_map.get("type") or col_map.get("vendor type")
    city_c = col_map.get("city")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        sup = normalize_supplier_code(ws.cell(r, sup_c).value if sup_c else "")
        if not sup:
            continue
        name = normalize_text(ws.cell(r, name_c).value if name_c else "").upper()
        group = normalize_text(ws.cell(r, group_c).value if group_c else "")
        view = normalize_text(ws.cell(r, view_c).value if view_c else "")
        supplier_type = normalize_text(ws.cell(r, type_c).value if type_c else "")
        city = normalize_text(ws.cell(r, city_c).value if city_c else "").upper()
        if not supplier_type:
            supplier_type = group
        rows.append({
            "Supplier": sup,
            "Master Name": name,
            "Supplier Type": supplier_type,
            "Group": group,
            "View - Y/N": view,
            "City": city,
        })

    master_df = pd.DataFrame(rows).drop_duplicates(subset=["Supplier"], keep="last")
    return master_df, wb, ws, col_map


EXPECTED_HEADERS = [
    "Supplier", "Vendor Name", "DocumentNo", "St", "Reference", "PK", "Type", "DD",
    "Doc. Date", "Pstng Date", "S", "Pmnt Date", "Local Crcy Amt", "LCu",
    "Amount in DC", "Crcy", "Text", "User Name", "Clearing"
]


def find_header_positions(raw_df):
    positions = []
    for idx in range(len(raw_df)):
        row_values = [normalize_text(v) for v in raw_df.iloc[idx].tolist()]
        lowered = [v.lower() for v in row_values]
        if "supplier" in lowered and "vendor name" in lowered and "documentno" in lowered:
            input_col_map = {}
            for col_no, val in enumerate(row_values):
                if val in EXPECTED_HEADERS:
                    input_col_map[val] = col_no
            if "Supplier" in input_col_map and "DocumentNo" in input_col_map:
                positions.append((idx, input_col_map))
    return positions


def is_subtotal_or_total_row(raw_row):
    first_cell = normalize_text(raw_row.iloc[0] if len(raw_row) > 0 else "")
    return first_cell in ("*", "**")


def get_cell(raw_row, col_map, col_name):
    idx = col_map.get(col_name)
    if idx is None or idx < 0 or idx >= len(raw_row):
        return ""
    return raw_row.iloc[idx]


def read_sap_os_from_upload(os_file, sheet_name=None):
    raw_df = pd.read_excel(os_file, sheet_name=sheet_name or 0, header=None, dtype=object)
    header_positions = find_header_positions(raw_df)
    if not header_positions:
        raise Exception("No SAP OS header found. Expected columns: Supplier, Vendor Name, DocumentNo.")

    output_rows = []
    block_no = 0
    for pos_no, (header_row_idx, input_col_map) in enumerate(header_positions):
        block_no += 1
        next_header_idx = header_positions[pos_no + 1][0] if pos_no + 1 < len(header_positions) else len(raw_df)
        for ridx in range(header_row_idx + 1, next_header_idx):
            raw_row = raw_df.iloc[ridx]
            if is_subtotal_or_total_row(raw_row):
                continue
            supplier = normalize_supplier_code(get_cell(raw_row, input_col_map, "Supplier"))
            document_no = normalize_reference(get_cell(raw_row, input_col_map, "DocumentNo"))
            if not supplier or not document_no or not is_valid_supplier_code(supplier):
                continue

            vendor_name = normalize_text(get_cell(raw_row, input_col_map, "Vendor Name")).upper()
            pmnt_dt = parse_date_safe(get_cell(raw_row, input_col_map, "Pmnt Date"))
            text_value = normalize_text(get_cell(raw_row, input_col_map, "Text"))
            dd_value = normalize_text(get_cell(raw_row, input_col_map, "DD"))
            real_due = parse_real_due_dt(text_value, pmnt_dt, dd_value)
            lc_amount = clean_amount(get_cell(raw_row, input_col_map, "Local Crcy Amt"))

            output_rows.append({
                "BlockNo": block_no,
                "Supplier": supplier,
                "Name": vendor_name,
                "DocumentNo": document_no,
                "St": normalize_text(get_cell(raw_row, input_col_map, "St")),
                "Reference": normalize_reference(get_cell(raw_row, input_col_map, "Reference")),
                "PK": normalize_text(get_cell(raw_row, input_col_map, "PK")),
                "Type": normalize_text(get_cell(raw_row, input_col_map, "Type")),
                "DD": dd_value,
                "Doc. Date": parse_date_safe(get_cell(raw_row, input_col_map, "Doc. Date")),
                "Pstng Date": parse_date_safe(get_cell(raw_row, input_col_map, "Pstng Date")),
                "S": normalize_text(get_cell(raw_row, input_col_map, "S")),
                "Pmnt date": pmnt_dt,
                "RealDue_Dt": real_due,
                "Cash Flow": ageing_bucket(real_due),
                "LC amnt": lc_amount,
                "Payable Amount": abs(lc_amount),
                "LCu": normalize_text(get_cell(raw_row, input_col_map, "LCu")),
                "Amount in DC": clean_amount(get_cell(raw_row, input_col_map, "Amount in DC")),
                "Crcy": normalize_text(get_cell(raw_row, input_col_map, "Crcy")),
                "Text": text_value,
                "User Name": normalize_text(get_cell(raw_row, input_col_map, "User Name")),
                "Clearing": normalize_text(get_cell(raw_row, input_col_map, "Clearing")),
            })
    return pd.DataFrame(output_rows)


def enrich_with_master(os_df, master_df):
    df = os_df.merge(master_df, on="Supplier", how="left")
    df["Name"] = df["Name"].where(df["Name"].astype(str).str.strip() != "", df["Master Name"].fillna(""))
    for col in ["Supplier Type", "Group", "View - Y/N", "City"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("")
    df["Supplier Type"] = df["Supplier Type"].where(df["Supplier Type"].astype(str).str.strip() != "", df["Group"])
    df["Status"] = df["View - Y/N"].apply(lambda x: "Operative" if "op" in str(x).lower() and "non" not in str(x).lower() else ("Non Operative" if "non" in str(x).lower() else normalize_text(x)))
    desired = [
        "BlockNo", "Supplier", "Name", "Supplier Type", "Group", "View - Y/N", "Status", "City",
        "DocumentNo", "Reference", "Type", "DD", "Doc. Date", "Pstng Date", "Pmnt date", "RealDue_Dt",
        "Cash Flow", "LC amnt", "Payable Amount", "LCu", "Amount in DC", "Crcy", "Text", "User Name", "Clearing"
    ]
    remaining = [c for c in df.columns if c not in desired and c != "Master Name"]
    return df[[c for c in desired if c in df.columns] + remaining]


def pivot_report(df, index_cols, value_col="Payable Amount"):
    if df.empty:
        return pd.DataFrame()
    p = pd.pivot_table(
        df,
        values=value_col,
        index=index_cols,
        columns="Cash Flow",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    for bucket in BUCKET_ORDER:
        if bucket not in p.columns:
            p[bucket] = 0
    p["Grand Total"] = p[BUCKET_ORDER].sum(axis=1)
    p = p[index_cols + BUCKET_ORDER + ["Grand Total"]]
    return p.sort_values("Grand Total", ascending=False)


def style_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:2000]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)


def create_excel_download(final_df, supplier_report, group_report, status_report, missing_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl", date_format="DD-MM-YYYY", datetime_format="DD-MM-YYYY") as writer:
        final_df.to_excel(writer, index=False, sheet_name="SAP_OS_Clean")
        supplier_report.to_excel(writer, index=False, sheet_name="Supplier Ageing")
        group_report.to_excel(writer, index=False, sheet_name="Group Ageing")
        status_report.to_excel(writer, index=False, sheet_name="Op_NonOp")
        missing_df.to_excel(writer, index=False, sheet_name="Missing Suppliers")
        for ws in writer.book.worksheets:
            style_excel_sheet(ws)
    output.seek(0)
    return output.getvalue()


def metric_card(label, value, note=""):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ============================================================
# SIDEBAR UPLOAD
# ============================================================
with st.sidebar:
    st.markdown("## 📤 Upload Files")
    st.markdown("Upload both files extracted/maintained for SAP creditor ageing.")
    os_file = st.file_uploader("1) Upload SAP OS Excel", type=["xlsx", "xls"], help="SAP outstanding file exported from SAP")
    master_file = st.file_uploader("2) Upload Supplier Master Excel", type=["xlsx"], help="Master with Supplier, Name, Group, View - Y/N")
    st.markdown("---")
    use_positive_amount = st.toggle("Show payable as positive amount", value=True)
    st.caption("For creditors, SAP values may be negative. Dashboard uses payable amount as positive by default.")
    st.markdown("---")
    st.markdown("### Report Includes")
    st.markdown("✅ Supplier ageing  \n✅ Group ageing  \n✅ Operative vs Non-Operative  \n✅ Missing suppliers  \n✅ Excel download")

# ============================================================
# MAIN BODY
# ============================================================
if not os_file or not master_file:
    c1, c2 = st.columns([1.2, 1])
    with c1:
        st.markdown(
            """
            <div class="section-card">
                <h3>🚀 Creditor Ageing Web Application</h3>
                <p>This app converts your SAP creditor outstanding file into a clean ageing dashboard and downloadable Excel report.</p>
                <p><b>Required files:</b></p>
                <ol>
                    <li><b>SAP OS Excel</b> extracted from SAP</li>
                    <li><b>Supplier Master Excel</b> containing Supplier, Name, Group and View - Y/N</li>
                </ol>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            """
            <div class="section-card">
                <h3>📌 Output Focus</h3>
                <p>Supplier Name, Supplier Type, Group, Operative / Non-Operative and Amount ageing.</p>
                <p>Upload the files from the left sidebar to begin.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.stop()

try:
    with st.spinner("Processing SAP OS and Supplier Master files..."):
        master_df, master_wb, master_ws, master_col_map = read_master_from_upload(master_file)
        os_df = read_sap_os_from_upload(os_file)
        final_df = enrich_with_master(os_df, master_df)
        amount_col = "Payable Amount" if use_positive_amount else "LC amnt"

        missing_df = final_df[final_df["Group"].astype(str).str.strip().eq("")][["Supplier", "Name"]].drop_duplicates().copy()
        if not missing_df.empty:
            missing_df["Suggested Action"] = "Add / update supplier in master file"

        supplier_report = pivot_report(final_df, ["Supplier Type", "Group", "View - Y/N", "Supplier", "Name"], amount_col)
        group_report = pivot_report(final_df, ["Supplier Type", "Group"], amount_col)
        status_report = pivot_report(final_df, ["Status"], amount_col)
        excel_bytes = create_excel_download(final_df, supplier_report, group_report, status_report, missing_df)

    st.markdown('<div class="small-success">✅ Files processed successfully. Dashboard is ready.</div>', unsafe_allow_html=True)

    total_amount = final_df[amount_col].sum()
    overdue_amount = final_df.loc[final_df["Cash Flow"].eq("OVERDUE"), amount_col].sum()
    supplier_count = final_df["Supplier"].nunique()
    nonop_amount = final_df.loc[final_df["Status"].str.lower().str.contains("non", na=False), amount_col].sum()

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        metric_card("Total Outstanding", format_inr(total_amount), "Based on uploaded SAP OS")
    with m2:
        metric_card("Overdue Amount", format_inr(overdue_amount), "Real due date wise")
    with m3:
        metric_card("Total Suppliers", f"{supplier_count:,}", "Unique supplier codes")
    with m4:
        metric_card("Non-Operative Exposure", format_inr(nonop_amount), "As per View - Y/N master")

    st.markdown("### 🔎 Dashboard Filters")
    f1, f2, f3, f4 = st.columns(4)
    with f1:
        selected_status = st.multiselect("Operative Status", sorted([x for x in final_df["Status"].dropna().unique() if x]), default=[])
    with f2:
        selected_group = st.multiselect("Group", sorted([x for x in final_df["Group"].dropna().unique() if x]), default=[])
    with f3:
        selected_type = st.multiselect("Supplier Type", sorted([x for x in final_df["Supplier Type"].dropna().unique() if x]), default=[])
    with f4:
        selected_bucket = st.multiselect("Ageing Bucket", BUCKET_ORDER, default=[])

    filtered_df = final_df.copy()
    if selected_status:
        filtered_df = filtered_df[filtered_df["Status"].isin(selected_status)]
    if selected_group:
        filtered_df = filtered_df[filtered_df["Group"].isin(selected_group)]
    if selected_type:
        filtered_df = filtered_df[filtered_df["Supplier Type"].isin(selected_type)]
    if selected_bucket:
        filtered_df = filtered_df[filtered_df["Cash Flow"].isin(selected_bucket)]

    tabs = st.tabs(["📊 Overview", "👥 Supplier Ageing", "🏷️ Group Report", "✅ Operative / Non-Operative", "⚠️ Missing Master", "📄 Clean Data", "⬇️ Download"])

    with tabs[0]:
        c1, c2 = st.columns(2)
        with c1:
            bucket_summary = filtered_df.groupby("Cash Flow", as_index=False)[amount_col].sum()
            bucket_summary["Cash Flow"] = pd.Categorical(bucket_summary["Cash Flow"], BUCKET_ORDER, ordered=True)
            bucket_summary = bucket_summary.sort_values("Cash Flow")
            fig = px.bar(bucket_summary, x="Cash Flow", y=amount_col, title="Ageing Bucket Wise Outstanding", text_auto=".2s")
            fig.update_layout(height=430, margin=dict(l=10, r=10, t=55, b=10))
            st.plotly_chart(fig, width="stretch")
        with c2:
            status_summary = filtered_df.groupby("Status", as_index=False)[amount_col].sum().sort_values(amount_col, ascending=False)
            fig2 = px.pie(status_summary, names="Status", values=amount_col, title="Operative vs Non-Operative Amount")
            fig2.update_layout(height=430, margin=dict(l=10, r=10, t=55, b=10))
            st.plotly_chart(fig2, width="stretch")

        top_suppliers = filtered_df.groupby(["Supplier", "Name", "Group", "View - Y/N"], as_index=False)[amount_col].sum().sort_values(amount_col, ascending=False).head(15)
        fig3 = px.bar(top_suppliers, x=amount_col, y="Name", orientation="h", title="Top 15 Suppliers by Outstanding", hover_data=["Supplier", "Group", "View - Y/N"])
        fig3.update_layout(height=560, yaxis={"categoryorder": "total ascending"}, margin=dict(l=10, r=10, t=55, b=10))
        st.plotly_chart(fig3, width="stretch")

    with tabs[1]:
        st.markdown("#### Supplier wise ageing report")
        report = pivot_report(filtered_df, ["Supplier Type", "Group", "View - Y/N", "Supplier", "Name"], amount_col)
        st.dataframe(report, width="stretch", height=560)

    with tabs[2]:
        st.markdown("#### Group and supplier type wise ageing")
        report = pivot_report(filtered_df, ["Supplier Type", "Group"], amount_col)
        st.dataframe(report, width="stretch", height=560)

    with tabs[3]:
        st.markdown("#### Operative / Non-Operative ageing")
        report = pivot_report(filtered_df, ["Status"], amount_col)
        st.dataframe(report, width="stretch", height=360)
        detail = pivot_report(filtered_df, ["Status", "Supplier", "Name"], amount_col)
        st.dataframe(detail, width="stretch", height=500)

    with tabs[4]:
        if missing_df.empty:
            st.markdown('<div class="small-success">✅ No missing suppliers found in master.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="small-warning">⚠️ Some suppliers are missing / unmatched in the master file.</div>', unsafe_allow_html=True)
            st.dataframe(missing_df, width="stretch", height=420)

    with tabs[5]:
        st.markdown("#### Clean transaction data")
        st.dataframe(filtered_df, width="stretch", height=620)

    with tabs[6]:
        st.markdown("#### Download final creditor ageing report")
        st.download_button(
            label="⬇️ Download Complete Excel Report",
            data=excel_bytes,
            file_name=f"Tuareg_Creditor_Ageing_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
        csv_bytes = filtered_df.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            label="⬇️ Download Filtered Clean Data CSV",
            data=csv_bytes,
            file_name=f"Filtered_Creditor_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            width="stretch",
        )

except Exception as e:
    st.error("Processing failed. Please check the uploaded files and format.")
    st.exception(e)
